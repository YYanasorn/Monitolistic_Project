from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
from dataCollector import snapRun

def find_empty_column(ws, row):
    col = 3
    while ws.cell(row=row, column=col).value is not None:
        col += 1
    return get_column_letter(col)

def create_or_load_workbook():
    current_date = datetime.now().strftime("%Y-%m-%d")
    target_filename = f"{current_date}_Daily_Check_list.xlsx"
    if os.path.exists(target_filename):
        return load_workbook(target_filename)
    else:
        result_wb = load_workbook('result.xlsx')
        result_ws = result_wb.active
        result_wb.save(target_filename) 
        return load_workbook(target_filename) 

def FillWorkBook():
    wb = create_or_load_workbook()

    ip_addresses = {'A': '192.168.1.15', 'B': '192.168.1.12'}

    for device_letter, ip_address in ip_addresses.items():
        snap_data = snapRun(device_letter, ip_address)
        new_sheet_name = snap_data[0]

        if new_sheet_name not in wb.sheetnames:
            print(f"Sheet '{new_sheet_name}' not found, creating a new sheet...")
            new_sheet = wb.create_sheet(new_sheet_name, index=0)
        else:
            print(f"Sheet '{new_sheet_name}' found, using existing sheet...")
            new_sheet = wb[new_sheet_name]

        empty_column = find_empty_column(new_sheet, row=6)
        last_column_index = ord(empty_column) - 64

        for idx, param_value in enumerate(snap_data[1:], start=6):
            new_sheet.cell(row=idx, column=last_column_index, value=param_value)

    wb.save(f"{datetime.now().strftime('%Y-%m-%d')}_Daily_Check_list.xlsx")
    print("Saved")

FillWorkBook()
